from csv import DictReader
from openpyxl import load_workbook
import os

import yaml


# def read_csv(path):
#     with open(os.getcwd() + path, encoding="utf-8") as f:
#         reader = DictReader(f)
#         for d in reader:
#             yield d


def read_csv(path):
    with open(path, encoding="utf-8") as f:
        reader = DictReader(f)
        for d in reader:
            yield d


def read_xlsx(path):
    wb = load_workbook(path)
    ws = wb.active
    for c in ws.iter_rows(min_row=2, values_only=True):
        yield c


def read_yaml(key):
    with open(os.getcwd() + "/extract.yaml", encoding="utf-8", mode="r") as f:
        value = yaml.safe_load(f)
        return value[key]


def write_yaml(data):
    with open(os.getcwd() + "/extract.yaml", encoding="utf-8", mode="a+") as f:
        value = yaml.dump(data, stream=f, allow_unicode=True)


def clear_yaml():
    with open(os.getcwd() + "/extract.yaml", encoding="utf-8", mode="w") as f:
        f.truncate()


# safe_load 有问题，一般是对应的yaml的内容格式有问题
def read_testcase(testpath, path):
    a = os.getcwd()
    # 需要截取的路徑字段
    b = testpath
    if (a.index(b)):
        c = a.index(b)
        d = os.getcwd()[:c]
        print(os.getcwd()[:c])
        print(d + path)
    else:
        d = a
    with open(d + path, encoding="utf-8", mode="r") as f:
        value = yaml.safe_load(f)
        return value


if __name__ == "__main__":
    # for d in read_csv("../datas/login.csv"):
    #     print(d)
    # # print(r"D:\Learn\PYTHON\GitTest\tripo-test\Tripo\Tripo_Web\datas\login.csv")
    # for d in read_xlsx(r"D:\Learn\PYTHON\GitTest\tripo-test\Tripo\Tripo_Web\datas\login.xlsx"):
    #     print(d)
    # a=os.getcwd()
    # # c=os.getcwd().index("pages")
    # if(a.index("pages")):
    #     c=a.index("pages")
    #     print(os.getcwd()[:c])
    print(read_testcase("pages","datas\login.yaml"))
